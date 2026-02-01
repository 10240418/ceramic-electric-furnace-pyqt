"""
数据导出服务 - 导出历史数据为 Excel
"""
from datetime import datetime
from pathlib import Path
from loguru import logger
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class DataExporter:
    """数据导出服务"""
    
    # 1. 导出历史数据为 Excel
    @staticmethod
    def export_to_excel(
        file_path: str,
        batch_code: str,
        start_time: datetime,
        end_time: datetime,
        data_type: str,
        data_list: list
    ) -> bool:
        """
        导出数据为 Excel
        
        参数:
            file_path: 保存路径
            batch_code: 批次编号
            start_time: 起始时间
            end_time: 截止时间
            data_type: 数据类型（如 "1#弧流"）
            data_list: 数据列表，格式 [{"time": datetime, "value": float}, ...]
        
        返回:
            bool: 是否成功
        """
        try:
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "历史数据"
            
            # 设置列宽
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 22
            ws.column_dimensions['C'].width = 22
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 15
            
            # 表头样式
            header_font = Font(name='Microsoft YaHei', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # 写入表头
            headers = ['批次编号', '起始时间', '截止时间', '数据类型', '数据值']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # 数据样式
            data_font = Font(name='Microsoft YaHei', size=11)
            data_alignment = Alignment(horizontal='center', vertical='center')
            
            # 格式化时间
            start_time_str = start_time.strftime('%Y-%m-%d %H:%M:%S')
            end_time_str = end_time.strftime('%Y-%m-%d %H:%M:%S')
            
            # 写入数据
            for row_idx, data_point in enumerate(data_list, start=2):
                # 批次编号
                cell = ws.cell(row=row_idx, column=1, value=batch_code)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                # 起始时间
                cell = ws.cell(row=row_idx, column=2, value=start_time_str)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                # 截止时间
                cell = ws.cell(row=row_idx, column=3, value=end_time_str)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                # 数据类型
                cell = ws.cell(row=row_idx, column=4, value=data_type)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                # 数据值
                value = data_point.get('value', 0)
                cell = ws.cell(row=row_idx, column=5, value=round(value, 2))
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
            
            # 保存文件
            wb.save(file_path)
            logger.info(f"数据导出成功: {file_path}, 共 {len(data_list)} 条数据")
            return True
            
        except Exception as e:
            logger.error(f"导出数据失败: {e}", exc_info=True)
            return False


# 单例
_data_exporter = None


def get_data_exporter() -> DataExporter:
    """获取数据导出服务单例"""
    global _data_exporter
    if _data_exporter is None:
        _data_exporter = DataExporter()
    return _data_exporter

