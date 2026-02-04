"""
历史曲线页面 - 九宫格测试版本
"""
from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QGridLayout, QLabel, QHBoxLayout,
                              QScrollArea, QFileDialog, QMessageBox, QPushButton)
from PyQt6.QtCore import Qt, QTimer, QDateTime, QThread, pyqtSignal
from ui.styles.themes import ThemeManager
from ui.widgets.history_curve.chart_line import ChartLine
from ui.widgets.history_curve.bar_history import BarHistory
from ui.pages.page_batch_compare import PageBatchCompare
from backend.bridge.history_query import get_history_query_service
from datetime import datetime
from loguru import logger
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class HistoryQueryThread(QThread):
    """历史数据查询线程"""
    
    query_finished = pyqtSignal(dict)
    query_failed = pyqtSignal(str)
    
    def __init__(self, history_service, batch_code: str, start_time: datetime, end_time: datetime, interval: str):
        super().__init__()
        self.history_service = history_service
        self.batch_code = batch_code
        self.start_time = start_time
        self.end_time = end_time
        self.interval = interval
    
    def run(self):
        """在后台线程中执行查询"""
        try:
            logger.info(f"后台线程开始查询批次 {self.batch_code} 的历史数据")
            
            arc_data = self.history_service.query_arc_data(
                self.batch_code, self.start_time, self.end_time, self.interval
            )
            
            electrode_data = self.history_service.query_electrode_depth(
                self.batch_code, self.start_time, self.end_time, self.interval
            )
            
            power_data = self.history_service.query_power_energy(
                self.batch_code, self.start_time, self.end_time, self.interval
            )
            
            hopper_data = self.history_service.query_hopper_data(
                self.batch_code, self.start_time, self.end_time, self.interval
            )
            
            cooling_data = self.history_service.query_cooling_water(
                self.batch_code, self.start_time, self.end_time, self.interval
            )
            
            result = {
                "arc_data": arc_data,
                "electrode_data": electrode_data,
                "power_data": power_data,
                "hopper_data": hopper_data,
                "cooling_data": cooling_data
            }
            
            logger.info("后台线程查询完成，发送结果到主线程")
            self.query_finished.emit(result)
            
        except Exception as e:
            logger.error(f"后台线程查询失败: {e}", exc_info=True)
            self.query_failed.emit(str(e))


class PageHistoryCurve(QWidget):
    """历史曲线页面 - 九宫格测试版本"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.theme_manager = ThemeManager.instance()
        
        self.history_service = get_history_query_service()
        
        self.is_history_mode = False
        self.selected_batch = ""
        self.selected_batches = []
        
        self.current_start_time = None
        self.current_end_time = None
        
        self.cached_data = {
            "arc_current": {"1": [], "2": [], "3": []},
            "arc_voltage": {"1": [], "2": [], "3": []},
            "electrode_depth": {"1": [], "2": [], "3": []},
            "power": [],
            "energy": [],
            "feeding_total": [],
            "shell_water": [],
            "cover_water": []
        }
        
        self.init_ui()
        self.apply_styles()
        
        self.theme_manager.theme_changed.connect(self.on_theme_changed)
        
        self.is_initialized = False
        self.query_thread = None
        self.load_retry_count = 0
        self.max_retry_count = 10
    
    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(8, 8, 8, 8)
        main_layout.setSpacing(8)
        
        # 顶部控制栏
        self.bar_history = BarHistory()
        self.bar_history.mode_changed.connect(self.on_mode_changed)
        self.bar_history.batch_changed.connect(self.on_batch_changed)
        self.bar_history.batches_changed.connect(self.on_batches_changed)
        self.bar_history.time_range_changed.connect(self.on_time_range_changed)
        self.bar_history.query_clicked.connect(lambda: self.on_query_clicked(use_mock=False))
        self.bar_history.export_clicked.connect(self.on_export_all_data)
        self.bar_history.batch_time_range_loaded.connect(self.on_batch_time_range_loaded)
        main_layout.addWidget(self.bar_history)
        
        # 内容区域
        self.content_widget = QWidget()
        self.content_layout = QVBoxLayout(self.content_widget)
        self.content_layout.setContentsMargins(0, 0, 0, 0)
        self.content_layout.setSpacing(8)
        
        # 创建九宫格布局
        self.create_nine_grid_layout()
        
        # 批次对比页面
        self.batch_compare_page = PageBatchCompare()
        self.batch_compare_page.hide()
        self.content_layout.addWidget(self.batch_compare_page)
        
        # 滚动区域
        scroll_area = QScrollArea()
        scroll_area.setWidget(self.content_widget)
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        
        main_layout.addWidget(scroll_area)
    
    def create_nine_grid_layout(self):
        """创建 3x3 九宫格布局"""
        self.grid_container = QWidget()
        grid_layout = QGridLayout(self.grid_container)
        grid_layout.setContentsMargins(0, 0, 0, 0)
        grid_layout.setSpacing(8)
        
        colors = self.theme_manager.get_colors()
        
        # 第一行：弧流、弧压、电极高度
        self.chart_arc_current = self._create_chart_with_border(
            ChartLine(y_label="", accent_color=colors.GLOW_CYAN, show_grid=True),
            "弧流 (KA)"
        )
        grid_layout.addWidget(self.chart_arc_current, 0, 0)
        
        self.chart_arc_voltage = self._create_chart_with_border(
            ChartLine(y_label="", accent_color=colors.GLOW_CYAN, show_grid=True),
            "弧压 (V)"
        )
        grid_layout.addWidget(self.chart_arc_voltage, 0, 1)
        
        self.chart_electrode_depth = self._create_chart_with_border(
            ChartLine(y_label="", accent_color=colors.GLOW_CYAN, show_grid=True),
            "电极高度 (mm)"
        )
        grid_layout.addWidget(self.chart_electrode_depth, 0, 2)
        
        # 第二行：功率、能耗、累计投料量
        self.chart_power = self._create_chart_with_border(
            ChartLine(y_label="", accent_color=colors.GLOW_ORANGE, show_grid=True),
            "功率 (kW)"
        )
        grid_layout.addWidget(self.chart_power, 1, 0)
        
        self.chart_energy = self._create_chart_with_border(
            ChartLine(y_label="", accent_color=colors.GLOW_ORANGE, show_grid=True),
            "能耗 (kWh)"
        )
        grid_layout.addWidget(self.chart_energy, 1, 1)
        
        self.chart_feeding = self._create_chart_with_border(
            ChartLine(y_label="", accent_color=colors.GLOW_BLUE, show_grid=True),
            "累计投料量 (kg)"
        )
        grid_layout.addWidget(self.chart_feeding, 1, 2)
        
        # 第三行：炉皮冷却水、炉盖冷却水、预留空白
        self.chart_shell_water = self._create_chart_with_border(
            ChartLine(y_label="", accent_color=colors.GLOW_GREEN, show_grid=True),
            "炉皮冷却水 (m3)"
        )
        grid_layout.addWidget(self.chart_shell_water, 2, 0)
        
        self.chart_cover_water = self._create_chart_with_border(
            ChartLine(y_label="", accent_color=colors.GLOW_GREEN, show_grid=True),
            "炉盖冷却水 (m3)"
        )
        grid_layout.addWidget(self.chart_cover_water, 2, 1)
        
        # 预留空白位置
        self.placeholder_widget = QWidget()
        self.placeholder_widget.setMinimumHeight(250)
        tm = self.theme_manager
        self.placeholder_widget.setStyleSheet(f"""
            QWidget {{
                background: {tm.card_bg()};
                border: 1px solid {tm.border_glow()};
                border-radius: 6px;
            }}
        """)
        
        # 在预留位置添加提示文字
        placeholder_layout = QVBoxLayout(self.placeholder_widget)
        placeholder_label = QLabel("预留曲线位置")
        placeholder_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        placeholder_label.setStyleSheet(f"color: {tm.text_secondary()}; font-size: 16px;")
        placeholder_layout.addWidget(placeholder_label)
        
        grid_layout.addWidget(self.placeholder_widget, 2, 2)
        
        self.content_layout.addWidget(self.grid_container)
    
    def _create_chart_with_border(self, chart: ChartLine, title: str) -> QWidget:
        """创建带边框和标题的图表容器"""
        tm = self.theme_manager
        colors = self.theme_manager.get_colors()
        is_dark = self.theme_manager.is_dark_mode()
        
        # 创建外层容器（带边框）
        outer_container = QWidget()
        outer_container.setMinimumHeight(250)
        outer_container.setStyleSheet(f"""
            QWidget {{
                background: {tm.card_bg()};
                border: 1px solid {tm.border_glow()};
                border-radius: 6px;
                padding: 0px;
            }}
        """)
        
        # 创建内层容器（放置标题和图表）
        inner_container = QWidget()
        inner_container.setStyleSheet(f"""
            QWidget {{
                background: transparent;
                border: none;
                border-radius: 2px;
            }}
        """)
        
        # 外层布局
        outer_layout = QVBoxLayout(outer_container)
        outer_layout.setContentsMargins(6, 6, 6, 6)
        outer_layout.setSpacing(0)
        outer_layout.addWidget(inner_container)
        
        # 内层布局
        inner_layout = QVBoxLayout(inner_container)
        inner_layout.setContentsMargins(6, 6, 6, 6)
        inner_layout.setSpacing(4)
        
        # 创建标题栏（包含标题和复原按钮）
        title_bar = QWidget()
        title_bar.setStyleSheet("QWidget { background: transparent; border: none; }")
        title_bar_layout = QHBoxLayout(title_bar)
        title_bar_layout.setContentsMargins(0, 0, 0, 0)
        title_bar_layout.setSpacing(8)
        
        # 创建标题标签
        title_label = QLabel(title)
        title_label.setStyleSheet(f"""
            QLabel {{
                color: {tm.text_primary()};
                font-size: 14px;
                font-weight: bold;
                background: transparent;
                border: none;
                padding: 0px;
            }}
        """)
        title_label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        title_bar_layout.addWidget(title_label)
        
        title_bar_layout.addStretch()
        
        # 创建复原按钮（参考导出按钮样式）
        reset_btn = QPushButton("复原")
        reset_btn.setFixedHeight(24)
        reset_btn.setFixedWidth(60)
        reset_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        
        # 应用导出按钮样式
        if is_dark:
            bg_normal = f"rgba(42, 63, 95, 0.3)"
            bg_hover = f"rgba(42, 63, 95, 0.5)"
        else:
            bg_normal = tm.bg_light()
            bg_hover = tm.bg_medium()
        
        reset_btn.setStyleSheet(f"""
            QPushButton {{
                background: {bg_normal};
                color: {tm.text_primary()};
                border: 1px solid {tm.border_medium()};
                border-radius: 4px;
                padding: 4px 12px;
                font-size: 12px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: {bg_hover};
                border: 1px solid {tm.glow_cyan()};
            }}
            QPushButton:pressed {{
                background: {tm.bg_medium()};
            }}
        """)
        
        # 连接复原按钮点击事件
        reset_btn.clicked.connect(chart.reset_view)
        
        title_bar_layout.addWidget(reset_btn)
        
        # 添加标题栏和图表
        inner_layout.addWidget(title_bar)
        inner_layout.addWidget(chart)
        
        return outer_container
    
    def on_mode_changed(self, is_history_mode: bool):
        self.is_history_mode = is_history_mode
        
        if is_history_mode:
            self.grid_container.hide()
            self.batch_compare_page.show()
        else:
            self.batch_compare_page.hide()
            self.grid_container.show()
    
    def on_batch_changed(self, batch: str):
        logger.info(f"批次选择变化: {batch}")
        self.selected_batch = batch
        self.clear_cache()
    
    def on_batch_time_range_loaded(self, batch_code: str, start_time, end_time):
        logger.info(f"批次 {batch_code} 时间范围加载完成: {start_time} - {end_time}")
    
    def on_batches_changed(self, batches: list):
        self.selected_batches = batches
    
    def on_time_range_changed(self, start: QDateTime, end: QDateTime):
        logger.info(f"时间范围变化: {start.toString()} - {end.toString()}")
        self.clear_cache()
    
    def on_query_clicked(self, use_mock: bool = False):
        """查询按钮点击"""
        if self.is_history_mode:
            self.update_batch_compare(use_mock=use_mock)
        else:
            batch_code = self.bar_history.get_selected_batch()
            
            if not batch_code or batch_code in ["无数据", "加载失败"]:
                logger.warning("未选择有效批次")
                return
            
            start_qdt, end_qdt = self.bar_history.get_time_range()
            start_time = start_qdt.toPyDateTime()
            end_time = end_qdt.toPyDateTime()
            
            logger.info(f"查询按钮点击 - 批次: {batch_code}, 时间: {start_time} - {end_time}")
            
            self.query_all_data(batch_code, start_time, end_time)
    
    def query_all_data(self, batch_code: str, start_time: datetime, end_time: datetime):
        """查询所有数据"""
        if not batch_code:
            logger.warning("批次号为空，无法查询")
            return
        
        self.current_start_time = start_time
        self.current_end_time = end_time
        
        logger.info(f"开始查询批次 {batch_code} 的历史数据，时间范围: {start_time} - {end_time}")
        
        target_points = 100
        interval = self.history_service.calculate_optimal_interval(start_time, end_time, target_points=target_points)
        logger.info(f"使用聚合间隔: {interval}")
        
        if self.query_thread and self.query_thread.isRunning():
            logger.warning("上一次查询尚未完成，等待完成...")
            self.query_thread.wait()
        
        self.query_thread = HistoryQueryThread(
            self.history_service,
            batch_code,
            start_time,
            end_time,
            interval
        )
        
        self.query_thread.query_finished.connect(self.on_query_finished)
        self.query_thread.query_failed.connect(self.on_query_failed)
        
        self.query_thread.start()
        logger.info("后台查询线程已启动")
    
    def on_query_finished(self, result: dict):
        """查询完成回调"""
        try:
            logger.info("收到后台查询结果，开始更新缓存")
            
            arc_data = result.get("arc_data", {})
            self.cached_data["arc_current"] = arc_data.get("arc_current", {"U": [], "V": [], "W": []})
            self.cached_data["arc_voltage"] = arc_data.get("arc_voltage", {"U": [], "V": [], "W": []})
            
            self.cached_data["arc_current"]["1"] = self.cached_data["arc_current"].pop("U", [])
            self.cached_data["arc_current"]["2"] = self.cached_data["arc_current"].pop("V", [])
            self.cached_data["arc_current"]["3"] = self.cached_data["arc_current"].pop("W", [])
            
            self.cached_data["arc_voltage"]["1"] = self.cached_data["arc_voltage"].pop("U", [])
            self.cached_data["arc_voltage"]["2"] = self.cached_data["arc_voltage"].pop("V", [])
            self.cached_data["arc_voltage"]["3"] = self.cached_data["arc_voltage"].pop("W", [])
            
            electrode_data = result.get("electrode_data", {})
            self.cached_data["electrode_depth"] = electrode_data
            
            power_data = result.get("power_data", {})
            self.cached_data["power"] = power_data.get("power", [])
            self.cached_data["energy"] = power_data.get("energy", [])
            
            hopper_data = result.get("hopper_data", {})
            self.cached_data["feeding_total"] = hopper_data.get("feeding_total", [])
            
            cooling_data = result.get("cooling_data", {})
            self.cached_data["shell_water"] = cooling_data.get("shell_water", [])
            self.cached_data["cover_water"] = cooling_data.get("cover_water", [])
            
            logger.info("缓存更新完成，开始更新图表")
            
            self.update_all_charts()
            
        except Exception as e:
            logger.error(f"处理查询结果失败: {e}", exc_info=True)
    
    def on_query_failed(self, error_msg: str):
        """查询失败回调"""
        logger.error(f"历史数据查询失败: {error_msg}")
        QMessageBox.critical(self, "查询失败", f"查询历史数据失败:\n{error_msg}")
    
    def update_all_charts(self):
        """更新所有图表"""
        colors = self.theme_manager.get_colors()
        
        # 获取实际的图表对象（从容器中提取）
        chart_arc_current = self.chart_arc_current.findChild(ChartLine)
        chart_arc_voltage = self.chart_arc_voltage.findChild(ChartLine)
        chart_electrode_depth = self.chart_electrode_depth.findChild(ChartLine)
        chart_power = self.chart_power.findChild(ChartLine)
        chart_energy = self.chart_energy.findChild(ChartLine)
        chart_feeding = self.chart_feeding.findChild(ChartLine)
        chart_shell_water = self.chart_shell_water.findChild(ChartLine)
        chart_cover_water = self.chart_cover_water.findChild(ChartLine)
        
        # 1. 更新弧流图表
        data_1 = self.cached_data["arc_current"]["1"]
        data_2 = self.cached_data["arc_current"]["2"]
        data_3 = self.cached_data["arc_current"]["3"]
        
        if data_1 and data_2 and data_3:
            x_data = [self._parse_time_to_timestamp(d["time"]) for d in data_1]
            y_data_1 = [d["value"] for d in data_1]
            y_data_2 = [d["value"] for d in data_2]
            y_data_3 = [d["value"] for d in data_3]
            
            data_list = [
                ("1# (A)", x_data, y_data_1, colors.CHART_LINE_1),
                ("2# (A)", x_data, y_data_2, colors.CHART_LINE_2),
                ("3# (A)", x_data, y_data_3, colors.CHART_LINE_3),
            ]
            chart_arc_current.update_multi_data_with_time(data_list)
        
        # 2. 更新弧压图表
        data_1 = self.cached_data["arc_voltage"]["1"]
        data_2 = self.cached_data["arc_voltage"]["2"]
        data_3 = self.cached_data["arc_voltage"]["3"]
        
        if data_1 and data_2 and data_3:
            x_data = [self._parse_time_to_timestamp(d["time"]) for d in data_1]
            y_data_1 = [d["value"] for d in data_1]
            y_data_2 = [d["value"] for d in data_2]
            y_data_3 = [d["value"] for d in data_3]
            
            data_list = [
                ("1# (V)", x_data, y_data_1, colors.CHART_LINE_1),
                ("2# (V)", x_data, y_data_2, colors.CHART_LINE_2),
                ("3# (V)", x_data, y_data_3, colors.CHART_LINE_3),
            ]
            chart_arc_voltage.update_multi_data_with_time(data_list)
        
        # 3. 更新电极高度图表
        data_1 = self.cached_data["electrode_depth"]["1"]
        data_2 = self.cached_data["electrode_depth"]["2"]
        data_3 = self.cached_data["electrode_depth"]["3"]
        
        if data_1 and data_2 and data_3:
            x_data = [self._parse_time_to_timestamp(d["time"]) for d in data_1]
            y_data_1 = [d["value"] for d in data_1]
            y_data_2 = [d["value"] for d in data_2]
            y_data_3 = [d["value"] for d in data_3]
            
            data_list = [
                ("1# (mm)", x_data, y_data_1, colors.CHART_LINE_1),
                ("2# (mm)", x_data, y_data_2, colors.CHART_LINE_2),
                ("3# (mm)", x_data, y_data_3, colors.CHART_LINE_3),
            ]
            chart_electrode_depth.update_multi_data_with_time(data_list)
        
        # 4. 更新功率图表
        power_data = self.cached_data["power"]
        if power_data:
            x_data = [self._parse_time_to_timestamp(d["time"]) for d in power_data]
            y_data = [d["value"] for d in power_data]
            chart_power.update_data_with_time(x_data, y_data, "功率 (kW)")
        
        # 5. 更新能耗图表
        energy_data = self.cached_data["energy"]
        if energy_data:
            x_data = [self._parse_time_to_timestamp(d["time"]) for d in energy_data]
            y_data = [d["value"] for d in energy_data]
            chart_energy.update_data_with_time(x_data, y_data, "能耗 (kWh)")
        
        # 6. 更新累计投料量图表
        feeding_data = self.cached_data["feeding_total"]
        if feeding_data:
            x_data = [self._parse_time_to_timestamp(d["time"]) for d in feeding_data]
            y_data = [d["value"] for d in feeding_data]
            chart_feeding.update_data_with_time(x_data, y_data, "投料 (kg)")
        
        # 7. 更新炉皮冷却水图表
        shell_data = self.cached_data["shell_water"]
        if shell_data:
            x_data = [self._parse_time_to_timestamp(d["time"]) for d in shell_data]
            y_data = [d["value"] for d in shell_data]
            chart_shell_water.update_data_with_time(x_data, y_data, "炉皮 (m³)")
        
        # 8. 更新炉盖冷却水图表
        cover_data = self.cached_data["cover_water"]
        if cover_data:
            x_data = [self._parse_time_to_timestamp(d["time"]) for d in cover_data]
            y_data = [d["value"] for d in cover_data]
            chart_cover_water.update_data_with_time(x_data, y_data, "炉盖 (m³)")
    
    def _parse_time_to_timestamp(self, time_str: str) -> float:
        """将ISO格式时间字符串转换为时间戳"""
        from dateutil import parser
        dt = parser.isoparse(time_str)
        return dt.timestamp()
    
    def clear_cache(self):
        """清空数据缓存"""
        self.cached_data = {
            "arc_current": {"1": [], "2": [], "3": []},
            "arc_voltage": {"1": [], "2": [], "3": []},
            "electrode_depth": {"1": [], "2": [], "3": []},
            "power": [],
            "energy": [],
            "feeding_total": [],
            "shell_water": [],
            "cover_water": []
        }
        logger.info("数据缓存已清空")
    
    def update_batch_compare(self, use_mock: bool = False):
        """更新批次对比"""
        if not self.selected_batches:
            logger.warning("未选择批次，无法更新批次对比")
            return
        
        logger.info(f"开始查询批次对比数据: {self.selected_batches}")
        
        batch_data = {
            'batches': [],
            'energy_total': {},
            'feeding_total': {},
            'shell_water_total': {},
            'cover_water_total': {},
            'duration_hours': {}
        }
        
        if use_mock:
            logger.info("使用Mock模式，生成测试数据...")
            import random
            for batch_code in self.selected_batches:
                batch_data['batches'].append(batch_code)
                batch_data['energy_total'][batch_code] = random.uniform(15000, 35000)
                batch_data['feeding_total'][batch_code] = random.uniform(200000, 350000)
                batch_data['shell_water_total'][batch_code] = random.uniform(100, 350)
                batch_data['cover_water_total'][batch_code] = random.uniform(80, 180)
                batch_data['duration_hours'][batch_code] = random.uniform(12, 22)
        else:
            for batch_code in self.selected_batches:
                try:
                    stats = self.history_service.query_batch_statistics(batch_code)
                    
                    batch_data['batches'].append(batch_code)
                    batch_data['energy_total'][batch_code] = stats['energy_total']
                    batch_data['feeding_total'][batch_code] = stats['feeding_total']
                    batch_data['shell_water_total'][batch_code] = stats['shell_water_total']
                    batch_data['cover_water_total'][batch_code] = stats['cover_water_total']
                    batch_data['duration_hours'][batch_code] = stats['duration_hours']
                    
                except Exception as e:
                    logger.error(f"查询批次 {batch_code} 统计数据失败: {e}")
        
        self.batch_compare_page.update_batch_data(batch_data)
    
    def on_export_all_data(self):
        """导出数据（根据模式选择导出方式）"""
        if self.is_history_mode:
            # 批次对比模式：导出对比数据
            self.export_batch_compare_data()
        else:
            # 历史曲线模式：导出历史数据
            self.export_history_curve_data()
    
    def export_history_curve_data(self):
        """导出历史曲线数据到Excel文件，每个图表一个Sheet"""
        if not self.selected_batch:
            QMessageBox.warning(self, "导出失败", "请先选择批次并查询数据")
            return
        
        if not self.current_start_time or not self.current_end_time:
            QMessageBox.warning(self, "导出失败", "请先查询数据")
            return
        
        # 弹出文件保存对话框
        # 文件名格式：批次号_历史数据_起始时间_截止时间.xlsx
        start_str = self.current_start_time.strftime('%Y%m%d_%H%M%S')
        end_str = self.current_end_time.strftime('%Y%m%d_%H%M%S')
        default_filename = f"{self.selected_batch}_历史数据_{start_str}_{end_str}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出所有数据",
            default_filename,
            "Excel 文件 (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            # 创建工作簿
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # 删除默认的空白sheet
            
            # 表头样式
            header_font = Font(name='Microsoft YaHei', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # 数据样式
            data_font = Font(name='Microsoft YaHei', size=11)
            data_alignment = Alignment(horizontal='center', vertical='center')
            
            # 格式化时间
            start_time_str = self.current_start_time.strftime('%Y-%m-%d %H:%M:%S')
            end_time_str = self.current_end_time.strftime('%Y-%m-%d %H:%M:%S')
            
            # 定义要导出的数据集
            export_datasets = [
                ("弧流", self.cached_data["arc_current"], ["1#弧流 (A)", "2#弧流 (A)", "3#弧流 (A)"]),
                ("弧压", self.cached_data["arc_voltage"], ["1#弧压 (V)", "2#弧压 (V)", "3#弧压 (V)"]),
                ("电极高度", self.cached_data["electrode_depth"], ["1#电极 (mm)", "2#电极 (mm)", "3#电极 (mm)"]),
                ("功率", {"data": self.cached_data["power"]}, ["功率 (kW)"]),
                ("能耗", {"data": self.cached_data["energy"]}, ["能耗 (kWh)"]),
                ("累计投料量", {"data": self.cached_data["feeding_total"]}, ["投料 (kg)"]),
                ("炉皮冷却水", {"data": self.cached_data["shell_water"]}, ["炉皮 (m³)"]),
                ("炉盖冷却水", {"data": self.cached_data["cover_water"]}, ["炉盖 (m³)"]),
            ]
            
            # 为每个数据集创建一个Sheet
            for sheet_name, data_dict, labels in export_datasets:
                # 检查数据是否为空
                if isinstance(data_dict, dict):
                    if "data" in data_dict:
                        # 单曲线数据
                        data_list = data_dict["data"]
                        if not data_list:
                            logger.warning(f"{sheet_name} 数据为空，跳过")
                            continue
                        
                        # 创建工作表
                        ws = wb.create_sheet(title=sheet_name)
                        
                        # 设置列宽
                        ws.column_dimensions['A'].width = 20
                        ws.column_dimensions['B'].width = 22
                        ws.column_dimensions['C'].width = 22
                        ws.column_dimensions['D'].width = 20
                        ws.column_dimensions['E'].width = 15
                        
                        # 写入表头
                        headers = ['批次编号', '起始时间', '截止时间', '数据类型', '数据值']
                        for col_idx, header in enumerate(headers, start=1):
                            cell = ws.cell(row=1, column=col_idx, value=header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                            cell.border = border
                        
                        # 写入数据
                        for row_idx, data_point in enumerate(data_list, start=2):
                            ws.cell(row=row_idx, column=1, value=self.selected_batch).font = data_font
                            ws.cell(row=row_idx, column=1).alignment = data_alignment
                            ws.cell(row=row_idx, column=1).border = border
                            
                            ws.cell(row=row_idx, column=2, value=start_time_str).font = data_font
                            ws.cell(row=row_idx, column=2).alignment = data_alignment
                            ws.cell(row=row_idx, column=2).border = border
                            
                            ws.cell(row=row_idx, column=3, value=end_time_str).font = data_font
                            ws.cell(row=row_idx, column=3).alignment = data_alignment
                            ws.cell(row=row_idx, column=3).border = border
                            
                            ws.cell(row=row_idx, column=4, value=labels[0]).font = data_font
                            ws.cell(row=row_idx, column=4).alignment = data_alignment
                            ws.cell(row=row_idx, column=4).border = border
                            
                            value = data_point.get('value', 0)
                            ws.cell(row=row_idx, column=5, value=round(value, 2)).font = data_font
                            ws.cell(row=row_idx, column=5).alignment = data_alignment
                            ws.cell(row=row_idx, column=5).border = border
                    else:
                        # 三相数据
                        has_data = False
                        for phase_id in ["1", "2", "3"]:
                            if data_dict.get(phase_id):
                                has_data = True
                                break
                        
                        if not has_data:
                            logger.warning(f"{sheet_name} 数据为空，跳过")
                            continue
                        
                        # 创建工作表
                        ws = wb.create_sheet(title=sheet_name)
                        
                        # 设置列宽
                        ws.column_dimensions['A'].width = 20
                        ws.column_dimensions['B'].width = 22
                        ws.column_dimensions['C'].width = 22
                        ws.column_dimensions['D'].width = 20
                        ws.column_dimensions['E'].width = 15
                        
                        # 写入表头
                        headers = ['批次编号', '起始时间', '截止时间', '数据类型', '数据值']
                        for col_idx, header in enumerate(headers, start=1):
                            cell = ws.cell(row=1, column=col_idx, value=header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                            cell.border = border
                        
                        # 写入三相数据
                        row_idx = 2
                        for phase_idx, phase_id in enumerate(["1", "2", "3"]):
                            data_list = data_dict.get(phase_id, [])
                            if not data_list:
                                continue
                            
                            for data_point in data_list:
                                ws.cell(row=row_idx, column=1, value=self.selected_batch).font = data_font
                                ws.cell(row=row_idx, column=1).alignment = data_alignment
                                ws.cell(row=row_idx, column=1).border = border
                                
                                ws.cell(row=row_idx, column=2, value=start_time_str).font = data_font
                                ws.cell(row=row_idx, column=2).alignment = data_alignment
                                ws.cell(row=row_idx, column=2).border = border
                                
                                ws.cell(row=row_idx, column=3, value=end_time_str).font = data_font
                                ws.cell(row=row_idx, column=3).alignment = data_alignment
                                ws.cell(row=row_idx, column=3).border = border
                                
                                ws.cell(row=row_idx, column=4, value=labels[phase_idx]).font = data_font
                                ws.cell(row=row_idx, column=4).alignment = data_alignment
                                ws.cell(row=row_idx, column=4).border = border
                                
                                value = data_point.get('value', 0)
                                ws.cell(row=row_idx, column=5, value=round(value, 2)).font = data_font
                                ws.cell(row=row_idx, column=5).alignment = data_alignment
                                ws.cell(row=row_idx, column=5).border = border
                                
                                row_idx += 1
            
            # 检查是否创建了任何工作表
            if len(wb.sheetnames) == 0:
                QMessageBox.warning(self, "导出失败", "没有可导出的数据")
                return
            
            # 保存文件
            wb.save(file_path)
            logger.info(f"所有数据导出成功: {file_path}")
            QMessageBox.information(self, "导出成功", f"所有数据已导出到:\n{file_path}\n\n共 {len(wb.sheetnames)} 个数据表")
            
        except Exception as e:
            logger.error(f"导出数据失败: {e}", exc_info=True)
            QMessageBox.critical(self, "导出失败", f"导出数据时发生错误:\n{str(e)}")
    
    def export_batch_compare_data(self):
        """导出批次对比数据到Excel文件（单个Sheet）"""
        if not self.selected_batches:
            QMessageBox.warning(self, "导出失败", "请先选择批次并查询数据")
            return
        
        # 获取批次对比数据
        batch_data = self.batch_compare_page.batch_compare_widget.batch_data
        if not batch_data or not batch_data.get('batches'):
            QMessageBox.warning(self, "导出失败", "没有可导出的数据")
            return
        
        # 弹出文件保存对话框
        # 文件名格式：累计数据对比_导出时间.xlsx
        export_time_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"累计数据对比_{export_time_str}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出批次对比数据",
            default_filename,
            "Excel 文件 (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "批次对比数据"
            
            # 设置列宽
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 22
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            
            # 表头样式
            header_font = Font(name='Microsoft YaHei', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # 数据样式
            data_font = Font(name='Microsoft YaHei', size=11)
            data_alignment = Alignment(horizontal='center', vertical='center')
            
            # 写入表头
            headers = ['批次号', '时间', '数据类型', '数据值']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # 获取当前时间作为导出时间
            export_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # 定义数据类型列表（按顺序）
            data_types = [
                ('累计能耗', 'energy_total', 'kWh'),
                ('累计投料', 'feeding_total', 'kg'),
                ('炉皮累计水流量', 'shell_water_total', 'm³'),
                ('炉盖累计水流量', 'cover_water_total', 'm³'),
                ('开炉时长', 'duration_hours', '小时')
            ]
            
            # 写入数据
            row_idx = 2
            batches = batch_data.get('batches', [])
            
            # 按数据类型分组，每个数据类型显示所有批次的数据
            for data_type_name, data_key, unit in data_types:
                data_dict = batch_data.get(data_key, {})
                
                # 为每个批次写入该数据类型的值
                for batch_code in batches:
                    value = data_dict.get(batch_code, 0)
                    
                    # 批次号
                    cell = ws.cell(row=row_idx, column=1, value=batch_code)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = border
                    
                    # 时间
                    cell = ws.cell(row=row_idx, column=2, value=export_time)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = border
                    
                    # 数据类型
                    cell = ws.cell(row=row_idx, column=3, value=f"{data_type_name} ({unit})")
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = border
                    
                    # 数据值
                    cell = ws.cell(row=row_idx, column=4, value=round(value, 2))
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = border
                    
                    row_idx += 1
            
            # 保存文件
            wb.save(file_path)
            logger.info(f"批次对比数据导出成功: {file_path}")
            QMessageBox.information(
                self, 
                "导出成功", 
                f"批次对比数据已导出到:\n{file_path}\n\n共 {len(batches)} 个批次，{len(data_types)} 种数据类型"
            )
            
        except Exception as e:
            logger.error(f"导出批次对比数据失败: {e}", exc_info=True)
            QMessageBox.critical(self, "导出失败", f"导出数据时发生错误:\n{str(e)}")
    
    def apply_styles(self):
        tm = self.theme_manager
        
        self.setStyleSheet(f"""
            PageHistoryCurve {{
                background: {tm.bg_deep()};
            }}
            QScrollArea {{
                background: {tm.bg_deep()};
                border: none;
            }}
            QWidget {{
                background: {tm.bg_deep()};
            }}
            QScrollBar:vertical {{
                background: {tm.bg_medium()};
                width: 12px;
                border-radius: 6px;
            }}
            QScrollBar::handle:vertical {{
                background: {tm.border_medium()};
                border-radius: 6px;
                min-height: 30px;
            }}
            QScrollBar::handle:vertical:hover {{
                background: {tm.glow_cyan()};
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
            }}
        """)
    
    def on_theme_changed(self):
        self.apply_styles()
        self._refresh_chart_containers()
        self._refresh_reset_buttons()
    
    def _refresh_chart_containers(self):
        """刷新所有图表容器的样式"""
        tm = self.theme_manager
        
        # 刷新九宫格中的所有图表容器
        chart_containers = [
            self.chart_arc_current, self.chart_arc_voltage, self.chart_electrode_depth,
            self.chart_power, self.chart_energy, self.chart_feeding,
            self.chart_shell_water, self.chart_cover_water, self.placeholder_widget
        ]
        
        for container in chart_containers:
            if container:
                # 更新外层容器样式
                container.setStyleSheet(f"""
                    QWidget {{
                        background: {tm.card_bg()};
                        border: 1px solid {tm.border_glow()};
                        border-radius: 6px;
                        padding: 0px;
                    }}
                """)
                
                # 更新内部标题标签
                for label in container.findChildren(QLabel):
                    # 标题标签
                    if label.font().bold():
                        label.setStyleSheet(f"""
                            QLabel {{
                                color: {tm.text_primary()};
                                font-size: 14px;
                                font-weight: bold;
                                background: transparent;
                                border: none;
                                padding: 0px;
                            }}
                        """)
                    # 预留位置标签
                    elif label.text() == "预留曲线位置":
                        label.setStyleSheet(f"color: {tm.text_secondary()}; font-size: 16px;")
    
    def _refresh_reset_buttons(self):
        """刷新所有复原按钮的样式"""
        tm = self.theme_manager
        is_dark = self.theme_manager.is_dark_mode()
        
        if is_dark:
            bg_normal = f"rgba(42, 63, 95, 0.3)"
            bg_hover = f"rgba(42, 63, 95, 0.5)"
        else:
            bg_normal = tm.bg_light()
            bg_hover = tm.bg_medium()
        
        # 查找所有复原按钮并更新样式
        for btn in self.findChildren(QPushButton):
            if btn.text() == "复原":
                btn.setStyleSheet(f"""
                    QPushButton {{
                        background: {bg_normal};
                        color: {tm.text_primary()};
                        border: 1px solid {tm.border_medium()};
                        border-radius: 4px;
                        padding: 4px 12px;
                        font-size: 12px;
                        font-weight: bold;
                    }}
                    QPushButton:hover {{
                        background: {bg_hover};
                        border: 1px solid {tm.glow_cyan()};
                    }}
                    QPushButton:pressed {{
                        background: {tm.bg_medium()};
                    }}
                """)
    
    def showEvent(self, event):
        """页面显示时重新加载批次列表"""
        super().showEvent(event)
        
        use_mock = False
        
        if not self.is_initialized:
            self.load_retry_count = 0
            if use_mock:
                self.bar_history.load_batch_list(use_mock=True)
                self.is_initialized = True
            else:
                QTimer.singleShot(500, self.load_initial_data)
                self.is_initialized = True
        else:
            logger.info("页面显示，重新加载批次列表...")
            self.bar_history.load_batch_list(use_mock=use_mock)
    
    def load_initial_data(self):
        """初始加载数据"""
        try:
            logger.info("开始初始加载批次列表...")
            self.bar_history.load_batch_list(use_mock=False)
            logger.info("批次列表加载成功")
        except Exception as e:
            logger.error(f"初始加载失败: {e}")
            self.load_retry_count += 1
            if self.load_retry_count < self.max_retry_count:
                logger.info(f"将在 1 秒后重试 ({self.load_retry_count}/{self.max_retry_count})...")
                QTimer.singleShot(1000, self.load_initial_data)
            else:
                logger.error("达到最大重试次数，加载失败")
    
    def hideEvent(self, event):
        """页面隐藏时停止查询线程"""
        super().hideEvent(event)
        
        if self.query_thread and self.query_thread.isRunning():
            logger.info("页面隐藏，停止查询线程...")
            self.query_thread.quit()
            self.query_thread.wait()

